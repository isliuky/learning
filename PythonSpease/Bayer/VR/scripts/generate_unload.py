import re
from openpyxl import load_workbook


def extract_table_name(sql_query):
    # 使用正则表达式从SQL查询中提取表名
    match = re.search(r'from\s+(\w+\.\w+)', sql_query, re.IGNORECASE)
    if match:
        table_name = match.group(1).split('.')[-1]  # 只取点号后面的表名部分
        return table_name
    return None


def generate_unload_commands(excel_file, sheet_name, s3_bucket, iam_role, compression='GZIP'):
    wb = load_workbook(filename=excel_file)
    ws = wb[sheet_name]

    commands = []
    for row in ws.iter_rows(min_row=2):  # 跳过标题行
        # 确保每个单元格的值都存在，再调用 strip()
        query = " ".join([str(cell.value).strip() for cell in row if cell.value is not None])
        # 替换单引号为双美元符号
        query = query.replace("'", "$$")
        table_name = extract_table_name(query)
        if table_name:
            s3_path = f's3://{s3_bucket}/{table_name}.csv'
            unload_cmd = f"""UNLOAD ('{query}')
                            TO '{s3_path}'
                            IAM_ROLE '{iam_role}'
                            PARALLEL OFF
                            DELIMITER '|'
                            HEADER
                            {compression};"""
            commands.append(unload_cmd)

    return commands

# 使用函数
excel_file = r'C:\Users\EJOPV\OneDrive - Bayer\Desktop\新建 Microsoft Excel 工作表.xlsx'
sheet_name = 'Sheet2'
s3_bucket = 'ph-cdp-nprod-dev-cn-north-1/cn_cdp_vr/data_output'
iam_role = 'arn:aws-cn:iam::268754486553:role/ph-cdp-dev-redshift-spectrum'

# 生成压缩的CSV文件
commands = generate_unload_commands(excel_file, sheet_name, s3_bucket, iam_role, compression='GZIP')

for cmd in commands:
    print(cmd)